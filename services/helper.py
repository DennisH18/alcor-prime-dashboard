import streamlit as st
import pandas as pd
import calendar
import dropbox
from io import BytesIO
from collections import OrderedDict
import unicodedata
import re

import services.dropboxAuth as dropboxAuth
from services.supabaseService import supabase_client

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, PatternFill
import services.supabaseService as supabaseService

access_token = dropboxAuth.get_access_token()
dbx = dropbox.Dropbox(access_token)


@st.cache_data
def fetch_dropbox_data():
    data_store = {}
    try:
        for entry in dbx.files_list_folder("").entries:
            if isinstance(entry, dropbox.files.FolderMetadata):
                company = entry.name
                data_store[company] = {}

                try:
                    year_entries = dbx.files_list_folder(f"/{company}").entries
                except dropbox.exceptions.ApiError as e:
                    st.warning(f"Skipping {company} due to API error: {e}")
                    continue

                for year_entry in year_entries:
                    if (
                        isinstance(year_entry, dropbox.files.FolderMetadata)
                        and year_entry.name.isdigit()
                    ):
                        year = int(year_entry.name)
                        data_store[company][year] = {}

                        try:
                            file_entries = dbx.files_list_folder(
                                f"/{company}/{year}"
                            ).entries
                        except dropbox.exceptions.ApiError as e:
                            st.warning(
                                f"Skipping {company}/{year} due to API error: {e}"
                            )
                            continue

                        for file_entry in file_entries:

                            if isinstance(
                                file_entry, dropbox.files.FileMetadata
                            ) and file_entry.name.endswith(".xlsx"):
                                file_path = f"/{company}/{year}/{file_entry.name}"

                                try:
                                    _, res = dbx.files_download(file_path)
                                    xls = pd.ExcelFile(BytesIO(res.content))

                                    if "Management Report" in file_entry.name:
                                        df = pd.read_excel(xls, engine="openpyxl")
                                        data_store[company][year][file_entry.name] = df
                                    elif "Budget" in file_entry.name:
                                        df = pd.read_excel(
                                            xls, sheet_name=company, engine="openpyxl"
                                        )
                                        data_store[company][year][file_entry.name] = df
                                    elif "JPCC vs Others" in file_entry.name:
                                        jpcc_sheet_name = next(
                                            (
                                                s
                                                for s in xls.sheet_names
                                                if "jpcc vs others" in s.lower()
                                            ),
                                            None,
                                        )
                                        if jpcc_sheet_name:
                                            df = pd.read_excel(
                                                xls,
                                                sheet_name=jpcc_sheet_name,
                                                engine="openpyxl",
                                            )
                                            data_store[company][year][
                                                file_entry.name
                                            ] = df

                                except Exception as e:
                                    st.error(f"Error reading {file_entry.name}: {e}")

    except dropbox.exceptions.ApiError as e:
        st.error(f"Dropbox Error: {e}")

    return data_store


@st.cache_data
def get_available_months(data, companies, selected_year):
    months = set()

    for key, value in data.items():
        parts = key.split("_")
        if len(parts) == 3:
            company, month, year = parts
            if company in companies and year == str(selected_year):
                months.add(month)

    return (
        sorted(months, key=lambda m: list(calendar.month_abbr).index(m))
        if months
        else []
    )


@st.cache_data
def get_available_companies_and_years(data_store):
    companies = sorted(data_store.keys())
    years = sorted({year for company in data_store for year in data_store[company]})
    return companies, years[1:] if len(years) > 1 else []



def transform_to_category_codes(pnl_account_categories_dict):

    desired_order = [
        ("REVENUE", [
            "Revenue - Products", 
            "Revenues - Food and Beverage", 
            "Service & Rental", 
            "Other Revenue", 
            "Sales Discount & Return", 
        ]),
        ("COGS", [
            "COGS - Products", 
            "COGS - Food & Beverages", 
            "COGS - Services & Rental", 
            "Royalty", 
            "Others", 
        ]),
        ("HUMAN RESOURCES", [
            "Salary and Benefit",
            "Medical",
            "Other",
            "Casual Expense",
            "Freelance Expense",
            "Outsourcing Expense",
        ]),
        ("OPERATIONAL EXPENSES", [
            "Event & Promotion",
            "Other Marketing and Sales Expenses",
            "Utility Expense",
            "Telephone & Internet Expense",
            "Supplies Expense",
            "Rental Expense",
            "Travelling Expense",
            "Complimentary Expense",
            "Training & Seminars Expense",
            "Transportation Expense"
            "General Affair Expenses",
            "PPE Tax & Insurance Expense",
            "Other General Affair Expense",
        ]),
        ("DEPRECIATION & MAINTENANCE", [
            "Depr - Buildings",
            "Depr - Vehicle",
            "Depr - Furniture & Fixture",
            "Depr - Equipment",
            "Depr - Intangible Assets",
            "Depr - Leasehold Improvement",
            "Repair Maintenance"
        ]),
        ("OTHER INCOME / EXPENSES", [
            "Other Income",
            "Other Expenses",
            "Provision For Income Tax"
        ]),
    ]

    transformed_dict = OrderedDict()
    raw_transformed = {}

    for main_category, subcategories in pnl_account_categories_dict.items():
        if isinstance(subcategories, dict):
            raw_transformed[main_category] = subcategories

    used_codes = set()

    for main_category, subcategory_order in desired_order:
        if main_category not in raw_transformed:
            continue

        subcategories = raw_transformed[main_category]

        for subcat in subcategory_order:
            if subcat in subcategories:
                codes = subcategories[subcat]

                if main_category not in transformed_dict:
                    transformed_dict[main_category] = []

                transformed_dict[main_category].extend(codes)
                used_codes.update(codes)

    for main_category, subcategories in raw_transformed.items():
        if main_category not in transformed_dict:
            transformed_dict[main_category] = []

        for subcat, codes in subcategories.items():
            for code in codes:
                if code not in used_codes:
                    transformed_dict[main_category].append(code)
                    used_codes.add(code)

    transformed_items = list(transformed_dict.items())
    if "GROSS PROFIT" not in transformed_dict:
        transformed_items.insert(2, ("GROSS PROFIT", None))
    if "TOTAL EXPENSES" not in transformed_dict:
        transformed_items.insert(6, ("TOTAL EXPENSES", None))
    if "NET PROFIT" not in transformed_dict:
        transformed_items.append(("NET PROFIT", None))

    return OrderedDict(transformed_items)


def export_all_tables_to_excel(company_html_dict):
    output = BytesIO()
    wb = Workbook()

    header_fill = PatternFill(
        start_color="000000", end_color="000000", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")

    number_format = "#,##0;(#,##0)"
    percent_format = "0.0%;(0.0%)"

    for company, table_html in company_html_dict.items():
        soup = BeautifulSoup(table_html, "html.parser")
        table = soup.find("table")

        if table:
            ws = wb.create_sheet(title=company[:31])
            row_map = {}

            for row_idx, row in enumerate(table.find_all("tr"), start=1):
                col_idx = 1

                for cell in row.find_all(["td", "th"]):
                    while (row_idx, col_idx) in row_map:
                        col_idx += 1

                    colspan = int(cell.get("colspan", 1))
                    rowspan = int(cell.get("rowspan", 1))
                    value = cell.get_text(strip=True)

                    is_bold = (
                        cell.find("b")
                        or cell.find("strong")
                        or "font-weight: bold" in str(cell.get("style", "")).lower()
                    )

                    num_format = None

                    if value.endswith("%"):
                        try:
                            value = float(value.replace("%", "")) / 100
                            num_format = percent_format
                        except ValueError:
                            pass
                    elif "(" in value and ")" in value:
                        try:
                            value = -float(
                                value.replace("(", "").replace(")", "").replace(",", "")
                            )
                            num_format = number_format
                        except ValueError:
                            pass
                    elif "," in value:
                        try:
                            value = float(value.replace(",", ""))
                            num_format = number_format
                        except ValueError:
                            pass

                    cell_ref = ws.cell(row=row_idx, column=col_idx, value=value)

                    # Apply styles
                    if cell.name == "th":
                        cell_ref.fill = header_fill
                        cell_ref.font = header_font
                    elif is_bold:
                        cell_ref.font = Font(bold=True)

                    # Apply number format only if num_format is set
                    if num_format:
                        cell_ref.number_format = num_format

                    # Merge columns
                    if colspan > 1:
                        end_col = col_idx + colspan - 1
                        ws.merge_cells(
                            start_row=row_idx,
                            start_column=col_idx,
                            end_row=row_idx,
                            end_column=end_col,
                        )
                        col_idx = end_col

                    # Merge rows
                    if rowspan > 1:
                        for r in range(row_idx, row_idx + rowspan):
                            row_map[(r, col_idx)] = True
                        ws.merge_cells(
                            start_row=row_idx,
                            start_column=col_idx,
                            end_row=row_idx + rowspan - 1,
                            end_column=col_idx,
                        )

                    col_idx += 1

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(output)
    output.seek(0)
    return output


def get_pnl_account_categories_dict():

    df = pd.DataFrame(supabaseService.fetch_data("COA"))

    pnl_account_categories_dict = {}
    for _, row in df.iterrows():
        main_category = row["main_category"]
        subcategory = row["subcategory"]
        coa = row["coa"]
        description = row["description"]

        if main_category not in pnl_account_categories_dict:
            pnl_account_categories_dict[main_category] = {}
        if subcategory not in pnl_account_categories_dict[main_category]:
            pnl_account_categories_dict[main_category][subcategory] = {}
        pnl_account_categories_dict[main_category][subcategory][coa] = description

    return pnl_account_categories_dict


def get_all_coa():
    codes = []

    def recurse_dict(d):
        for key, value in d.items():
            if isinstance(value, dict):
                recurse_dict(value)
            else:
                codes.append(key)

    recurse_dict(get_pnl_account_categories_dict())

    return codes


def verify_user():
    if "access_token" in st.session_state and st.session_state["access_token"]:
        token = st.session_state["access_token"]

        try:
            response = supabase_client.auth.get_user(token)
            if response and response.user:
                st.session_state["authenticated"] = True
                st.session_state["user_id"] = response.user.id

                st.sidebar.write(f"**{response.user.user_metadata['name']}**")
                st.sidebar.write("")

                users_data = supabaseService.fetch_data("Users")

                # implement Role based access

                st.sidebar.page_link("pages/1_Dashboard.py", label="Dashboard")
                st.sidebar.page_link("pages/2_PNL_Report.py", label="PNL Report")
                st.sidebar.page_link("pages/3_COA.py", label="COA")
                st.sidebar.page_link(
                    "pages/4_JPCC_vs_Others.py", label="JPCC vs Others"
                )
                st.sidebar.page_link("pages/5_Users.py", label="Users")

                # st.sidebar.write("")
                # st.sidebar.page_link("Logout", key="logout")

                return True

            else:
                st.error(f"You are not logged in to access this page. Please log in.")
                return False

        except Exception as e:
            st.error(f"Error verifying user: {e}")
            return False

    return False
